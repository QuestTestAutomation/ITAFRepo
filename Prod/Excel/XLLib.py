"""
This library has functions related to the read,write and  append operations on Excel sheets.
"""

from Prod.Utilities.ExceptionLib import *

#import ExceptionLib
import inspect
from openpyxl import *
import os
import time
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.compat import range
import inspect

class XLLib():
    def __init__(self):
        pass

    def pri(self):
        print("Hello World")
        print(inspect.stack()[0][3])

    def create_workbook(self,file, sheetname):
        try :

            sheetexists = 0
            if os.path.isfile(file):

                mywb = load_workbook(file)
                print("XL file exists")
            else:
                print("XL file does not exists")
                mywb = Workbook()
                time.sleep(5)

            sheet = mywb.active

            sheet.title = sheetname
            mywb.save(file)
            time.sleep(5)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def create_worksheets(self,file, sheetnames):
        try :
            sheetexists = 0
            i = 0
            if os.path.isfile(file):
                mywb = load_workbook(file)
                print("XL file exists")
                for sheetname in sheetnames:
                    mywb.create_sheet(index=i, title=sheetname)
                    mywb.save(file)
                    time.sleep(3)


            else:
                print("XL file does not exists")
                mywb = Workbook()
                time.sleep(10)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def print_XL_cell_values(self,file, sheetname, irow, icolumn):
        try :



            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            for row in range(1, mysheet.max_row):
                for col in range(1, mysheet.max_column):
                    if mysheet.cell(column=col, row=row).value != 'None':
                        print(mysheet.cell(column=col, row=row).value)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def get_xl_cell_values(self, file, sheetname, irow, icolumn):
        try :
            cellvalues = []
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            for row in range(1, mysheet.max_row):
                for col in range(1, mysheet.max_column):
                    if mysheet.cell(column=col, row=row).value != 'None':
                        print(mysheet.cell(column=col, row=row).value)
                        cellvalues.append(mysheet.cell(column=col, row=row).value)
            return cellvalues
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def get_xl_column_letter(self,file, sheetname, columnvalue):
        try:

            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            for col in range(1, mysheet.max_column):
                if (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper():
                    id = format(get_column_letter(col))
                    break
            return id
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def get_xl_column_index(self,file, sheetname, columnvalue):
        try :
            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            print("max column is " + str(mysheet.max_column))
            for col in range(1, mysheet.max_column):
                # print "***************"
                # print str((mysheet.cell(column=col, row=1).value)).upper()
                # print (str(columnvalue)).upper()
                # print (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper()

                if (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper():
                    id = col
                    break
            return id
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def get_xl_cell_value(self,file, sheetname, irow, icolumn):
        try :

            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            id = mysheet.cell(column=icolumn, row=irow).value
            return id
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])


    def get_xl_cell_value_using_column_header(self,file, sheetname, irow, columnheader):
        try :

            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            icolumn = XLLib.get_xl_column_index(XLLibp3, file, sheetname, columnheader)
            print("The Column is : " + str(icolumn))
            id = mysheet.cell(column=int(icolumn), row=irow).value
            id = str(id).strip()
            return id
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def set_xl_cell_value(self,file, sheetname, irow, icolumn, cellvalue):
        try :

            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            mysheet.cell(column=icolumn, row=irow).value = cellvalue
            mywb.save(file)
            time.sleep(5)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def set_xl_cell_value_using_column_header(self,file, sheetname, irow, columnheader, cellvalue):
        try :

            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            icolumn = XLLib.get_XL_column_index(file, sheetname, columnheader)
            mysheet.cell(column=int(icolumn), row=irow).value = cellvalue
            mywb.save(file)
            time.sleep(5)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def create_xl_header_lists(self,file, sheetname, lists):
        try :
            col = 1
            XLLib.Create_Workbook(file, sheetname)
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            for list in lists:
                mysheet.cell(column=int(col), row=1).value = list
                col = int(col) + int(1)
            mywb.save(file)
            time.sleep(5)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def add_xl_header_column(self,file, sheetname, columnheader):
        try :

            XLLib.Create_Workbook(file, sheetname)
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            col = int(mysheet.max_column) + int(1)
            mysheet.cell(column=int(col), row=1).value = columnheader
            mywb.save(file)
            time.sleep(5)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])


    def copy_xl_workbook(self,sourcefile, targetfile):

        try :

            if os.path.isfile(sourcefile):
                sourcewb = load_workbook(sourcefile)
                wslists = sourcewb.sheetnames

                for wslist in wslists:
                    print("sheet name is : " + str(wslist))
                    if os.path.isfile(targetfile):
                        mywb = load_workbook(targetfile)
                        print("XL file exists")
                    else:
                        print("XL file does not exists")
                        mywb = Workbook(targetfile)
                        # mywb.save(targetfile)
                        # time.sleep(5)

                    sourcews = sourcewb.get_sheet_by_name(str(wslist))
                    myws = sourcewb.create_sheet(index=int(sourcewb.get_index(sourcews)), title=str(wslist))
                    #   Myws = mywb.active
                    myws = sourcewb.copy_worksheet(sourcews)

                    mywb.save(targetfile)
                    time.sleep(5)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def get_xl_row_count(self,file, sheetname):
        try :
            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            id = mysheet.max_row
            return id
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def get_XL_column_count(self,file, sheetname):
        try :
            id = '0'
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            id = mysheet.max_column

            return id
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])


    def load_xl_cell_values_dictionary(self, file, sheetname):
        try :

            tabledict = {}
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]
            for row in range(1, mysheet.max_row):
                column = []
                for col in range(1, mysheet.max_column):
                    if mysheet.cell(column=col, row=row).value != 'None':
                        column.append(str(mysheet.cell(column=col, row=row).value))
                    else:
                        column.append('None')
                tabledict[str(row)] = column

            return tabledict
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])

    def get_table_row_as_dictionary(self,row,tabledictionary):

        rowdict = {}

        for index in range(len(tabledictionary[str(1)])):

            rowdict[tabledictionary[str(1)][index]] = tabledictionary[str(row)][index]
        print(rowdict)
        return rowdict

    def get_excel_row_as_dictionary(self,file,sheetname,row):
        tabledictionary = self.load_xl_cell_values_dictionary(self, file, sheetname)
        rowdict = self.get_table_row_as_dictionary(self,row,tabledictionary)

        return rowdict

    def get_xl_header_as_tuple(self, file, sheetname):
        try :

            columnlist[]
            mywb = load_workbook(file)
            mysheet = mywb[sheetname]

            for col in range(1, mysheet.max_column):
                if mysheet.cell(column=col, 1).value != 'None':
                    columnlist.append(str(mysheet.cell(column=col, 1).value))
                else:
                    columnlist.append('None')


            return tuple(columnlist)
        except Exception as e:
            HandleException(str(e), inspect.stack()[0][3])